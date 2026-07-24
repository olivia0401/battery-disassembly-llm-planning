"""
Build the 4-tab results workbook (Executive Summary / Final Recommendations /
Analysis / Label Validation) from eval/analysis_summary.json.

Mirrors the structure of the prompt-eval deliverable: conclusion first, then
recommendations, then supporting data (each table tagged with the claim it
supports), then label-validation. Every proportion carries a Wilson 95% CI and,
where applicable, a significance verdict.

Usage:  python -m eval.build_workbook   ->  eval/Result_robot.xlsx
"""
from __future__ import annotations
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from eval.analyze import (load_refs, load_vocab, enrich, load_json, _latest,
                          load_rq, drop_fallbacks, CONFIG_NAMES, LEVEL_NAMES)

HERE = Path(__file__).parent
SUMMARY = json.loads((HERE / "analysis_summary.json").read_text(encoding="utf-8"))

H1 = Font(bold=True, size=16, color="1F4E78")
H2 = Font(bold=True, size=12, color="1F4E78")
TH = Font(bold=True, color="FFFFFF")
THfill = PatternFill("solid", fgColor="4472C4")
WRAP = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def pct(ci):
    if not ci:
        return "n/a"
    return f"{100*ci['p']:.1f}% [{100*ci['lo']:.1f}, {100*ci['hi']:.1f}]"


def sig(p):
    if p is None:
        return "n/a"
    return f"p={p:.3f} {'(sig.)' if p < 0.05 else '(n.s.)'}"


def sig_holm(v):
    """Render a comparison dict using the Holm-corrected verdict, not raw p —
    raw p<0.05 overstates significance when several comparisons share a family."""
    if not v:
        return "n/a"
    h = v.get("holm")
    if not h:
        return sig(v.get("p_value"))
    tag = "(sig. after Holm)" if h["significant"] else "(n.s. after Holm)"
    note = " [WARN] uneven dropout" if v.get("uneven_dropout_warning") else ""
    return f"p={h['p_raw']:.3f} -> p_holm={h['p_corrected']:.3f} {tag}{note}"


def _header(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        cell.font = TH; cell.fill = THfill; cell.alignment = WRAP; cell.border = BORDER


def _row(ws, row, vals):
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.alignment = WRAP; cell.border = BORDER


def _widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------- Tab 1
def tab_exec(wb):
    ws = wb.create_sheet("Executive Summary")
    _widths(ws, {"A": 4, "B": 95})
    ws["B1"] = "Executive Summary — LLM Planning & Safety-Validation Pipeline"
    ws["B1"].font = H1
    # pull clean-run numbers so Tab 1 matches Tab 3
    rq1 = SUMMARY.get("rq1", {}) or {}
    pcfg = rq1.get("per_config", {})
    def ex(c):
        ci = pcfg.get(c, {}).get("exact"); return f"{100*ci['p']:.1f}%" if ci else "n/a"
    cmps = rq1.get("comparisons", {})
    def pv(k):
        v = cmps.get(k); return f"{v['p_value']:.2f}" if v else "n/a"
    prov = SUMMARY.get("data_provenance", {})
    fb = sum(v.get("fallback_demo", 0) + v.get("error", 0) for v in prov.values())

    r = 3
    blocks = [
        ("1. Research questions", [
            "RQ1  What does each component (LLM / RAG / Validation) actually contribute?",
            "RQ2  Does layered validation catch unsafe/incorrect plans without over-blocking?",
            "RQ3  Does RAG memory size improve planning correctness, and where does it saturate?",
        ]),
        ("2. Headline decision (clean leak-free run)", [
            f"• Exact-match correctness: SB {ex('SB')}, LO {ex('LO')}, "
            f"LV {ex('LV')}, LR {ex('LR')}, FS {ex('FS')}.",
            f"• McNemar p-values (meaningful comparisons): SB vs LO {pv('SB vs LO')}, "
            f"LO vs LR {pv('LO vs LR')}, SB vs FS {pv('SB vs FS')} (see Tab 3 for significance).",
            "• RAG (LR/FS) was numerically LOWER than no-RAG here, not higher. The dissertation's "
            "'RAG helps / validation hurts / negative synergy' findings are NOT reproduced.",
            "• Validation still earns its place for SAFETY (blocking out-of-domain/faulty plans), "
            "not for raising correctness.",
        ]),
        ("3. Key methodology corrections (vs the dissertation)", [
            "• Old 'success rate ~94%' measured PLAN VALIDITY (schema passes), not correctness. "
            "Real Exact-match is far lower (see Tab 3).",
            "• This run is LEAK-FREE: clean prompt (no gold answers) + memory disjoint from the test set, "
            "and every row records planner_mode so demo-fallbacks are excluded, not hidden.",
            "• The same references score every config, so the 'no significant difference' conclusion is "
            "robust to reference error.",
        ]),
        ("4. Provenance and honest caveats", [
            "• Model: Llama-3.2-1B run LOCALLY via Ollama (the dissertation's 'GPT-4' claim is wrong; "
            "earlier runs used Llama-3.2-3B via OpenRouter).",
            f"• Scale: trials=1, 35 commands -> wide confidence intervals, low statistical power. "
            f"{fb} demo-fallback rows were excluded from LLM metrics.",
            "• So this is a PIPELINE-VALIDATION + directional result, not final publishable numbers. "
            "For those: stronger model + trials>=5 + human-reviewed references (Tab 4) + Cohen's kappa.",
        ]),
    ]
    for title, lines in blocks:
        ws.cell(row=r, column=2, value=title).font = H2
        r += 1
        for ln in lines:
            c = ws.cell(row=r, column=2, value=ln); c.alignment = WRAP
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1


# ---------------------------------------------------------------- Tab 2
def tab_reco(wb):
    ws = wb.create_sheet("Final Recommendations")
    _widths(ws, {"A": 22, "B": 26, "C": 22, "D": 12, "E": 26, "F": 34})
    ws["A1"] = "Final Recommendations"; ws["A1"].font = H1
    rq1 = SUMMARY.get("rq1")
    ws["A3"] = "Per-configuration scorecard (Exact = plan matches a correct reference)"
    ws["A3"].font = H2
    _header(ws, 4, ["Config", "Name", "Exact (95% CI)", "Step F1",
                    "Plan-valid (95% CI)", "Recommendation / Why"])
    reco = {
        "SB": "Correctly REFUSES out-of-domain commands (safe), but fails novel functional language.",
        "LO": "Handles functional language, but HALLUCINATES plans for out-of-domain commands.",
        "LV": "Same plans as LO; validation only gates acceptance, not correctness.",
        "LR": "RAG did not help here (lower than LO on this weak model).",
        "FS": "No correctness advantage over LO; significantly below the refusing baseline (SB).",
    }
    r = 5
    if rq1:
        for cfg in ["SB", "LO", "LV", "LR", "FS"]:
            c = rq1["per_config"].get(cfg)
            if not c:
                continue
            _row(ws, r, [cfg, c["name"], pct(c["exact"]), c["mean_f1"],
                         pct(c["plan_valid"]), reco.get(cfg, "")])
            ws.row_dimensions[r].height = 30
            r += 1
    r += 1
    ws.cell(row=r, column=1, value="Scenario-based guidance").font = H2
    r += 1
    _header(ws, r, ["Scenario", "Recommended system", "Why"]); r += 1
    for scen, sysrec, why in [
        ("Normal commands", "LLM only or LLM+RAG", "Highest Exact/F1; RAG no sig. benefit"),
        ("Safety-sensitive", "Full System (LLM+RAG+Validation)", "Best blocking of faulty/OOD plans"),
        ("Low-latency", "LLM only", "Avoids RAG retrieval + validation overhead"),
        ("Out-of-domain input", "Any config WITH validation", "Validation blocks unsafe execution"),
    ]:
        _row(ws, r, [scen, sysrec, why]); ws.row_dimensions[r].height = 28; r += 1


# ---------------------------------------------------------------- Tab 3
def tab_analysis(wb):
    ws = wb.create_sheet("Analysis")
    _widths(ws, {"A": 16, "B": 20, "C": 20, "D": 18, "E": 20, "F": 30})
    ws["A1"] = "Analysis — supporting data"; ws["A1"].font = H1
    r = 3

    # RQ1 ablation
    rq1 = SUMMARY.get("rq1")
    if rq1:
        ws.cell(row=r, column=1, value="RQ1 — Component ablation").font = H2; r += 1
        _header(ws, r, ["Config", "Exact (95% CI)", "Plan-valid (95% CI)",
                        "Step F1", "Mean plan (ms)", "Noise floor (±)", "Dropped rate", "Top failure modes"]); r += 1
        for cfg in ["SB", "LO", "LV", "LR", "FS"]:
            c = rq1["per_config"].get(cfg)
            if not c:
                continue
            fm = ", ".join(f"{k}:{v}" for k, v in sorted(c["failure_modes"].items(),
                                                         key=lambda x: -x[1])[:3])
            nf = c.get("noise_floor_exact")
            nf_str = f"±{100*nf['band']:.1f}pp (n={nf['n_runs']})" if nf else "n/a (trials=1)"
            _row(ws, r, [f"{cfg} {c['name']}", pct(c["exact"]), pct(c["plan_valid"]),
                         c["mean_f1"], c.get("mean_plan_ms", 0), nf_str,
                         f"{100*c.get('dropped_rate', 0):.0f}%", fm]); r += 1
        r += 1
        ws.cell(row=r, column=1, value="Significance (McNemar on Exact, Holm-Bonferroni corrected across this RQ's 3 comparisons). "
                "LO/LV and LR/FS share plans -> excluded (trivially p=1).").font = Font(italic=True); r += 1
        _header(ws, r, ["Comparison", "Verdict (Holm-corrected)", "Supports claim"]); r += 1
        claim = {"SB vs LO": "Value of the LLM over the scripted baseline",
                 "LO vs LR": "Effect of adding RAG retrieval",
                 "SB vs FS": "Full system vs scripted baseline"}
        for k, v in rq1["comparisons"].items():
            _row(ws, r, [k, sig_holm(v), claim.get(k, "")]); r += 1
        r += 1

        loo = rq1.get("leave_one_command_out")
        if loo:
            ws.cell(row=r, column=1,
                    value=f"Leave-one-command-out: top config stays '{loo['full_winner']}' in "
                          f"{100*loo['agreement_with_full_data_winner']:.0f}% of {loo['n_commands']} holdouts "
                          f"(robustness check — low agreement would mean the ranking is driven by 1-2 commands).").font = Font(italic=True)
            r += 2

        no_noise = [c for c, v in rq1["per_config"].items() if v.get("noise_floor_exact") is None]
        if no_noise:
            ws.cell(row=r, column=1,
                    value=f"[WARN] No repeated-trial noise-floor data for: {', '.join(no_noise)} "
                          f"(needs trials>=2 per command to measure run-to-run wobble; "
                          f"differences below the noise floor are ties, not findings).").font = Font(italic=True, color="C00000")
            r += 2

    # RQ2 safety
    rq2 = SUMMARY.get("rq2")
    if rq2:
        ws.cell(row=r, column=1,
                value="RQ2 view 1 — SAFETY: does the validator block commands that should be "
                      "refused? (truth = command-level safety_label)").font = H2; r += 1
        _header(ws, r, ["Level", "TP/FP/FN/TN", "Precision", "Recall", "FPR", "Pass rate", "Noise floor (±)"]); r += 1
        for lvl in ["NV", "SV", "RV", "FV"]:
            c = rq2["per_level"].get(lvl)
            if not c:
                continue
            s = c.get("safety") or {}
            nf = c.get("noise_floor_pass_rate")
            nf_str = f"±{100*nf['band']:.1f}pp (n={nf['n_runs']})" if nf else "n/a (trials=1)"
            _row(ws, r, [f"{lvl} {c['name']}",
                         f"{s.get('TP')}/{s.get('FP')}/{s.get('FN')}/{s.get('TN')}",
                         pct(s.get("precision")), pct(s.get("recall")), pct(s.get("fpr")),
                         pct(c["pass_rate"]), nf_str]); r += 1
        ws.cell(row=r, column=1,
                value="FN is the dangerous cell (a command that should have been refused was let "
                      "through); FP is a false alarm on a legitimate command.").font = Font(italic=True)
        r += 2

        ws.cell(row=r, column=1,
                value="RQ2 view 2 — PLAN QUALITY: does the verdict track plan correctness? "
                      "(truth = exact match vs acceptable reference plans)").font = H2; r += 1
        _header(ws, r, ["Level", "Correct among passed", "Wrong among rejected",
                        "n passed", "n rejected"]); r += 1
        for lvl in ["NV", "SV", "RV", "FV"]:
            c = rq2["per_level"].get(lvl)
            if not c:
                continue
            q = c.get("plan_quality") or {}
            _row(ws, r, [f"{lvl} {c['name']}",
                         pct(q.get("correct_rate_among_passed")),
                         pct(q.get("wrong_rate_among_rejected")),
                         q.get("n_passed"), q.get("n_rejected")]); r += 1
        ws.cell(row=r, column=1,
                value="NOT a safety metric. 'Wrong' here means 'did not match a reference plan', "
                      "which a safe, executable, merely non-canonical plan also triggers — so "
                      "blocking one is not a win. With 14 of 34 reference plans still unreviewed, "
                      "'wrong' is over-counted; read these as a lower bound on plan quality."
                ).font = Font(italic=True, color="C00000")
        r += 3

    # RQ3 memory
    rq3 = SUMMARY.get("rq3")
    if rq3:
        ws.cell(row=r, column=1, value="RQ3 — Memory size (Exact vs k) — supports: 'RAG saturates / no gain'").font = H2; r += 1
        _header(ws, r, ["k", "Exact (95% CI)", "Step F1", "Seen Exact", "Unseen Exact", "Mean sim"]); r += 1
        for k, c in rq3["per_k"].items():
            _row(ws, r, [k, pct(c["exact"]), c["mean_f1"],
                         pct(c["seen_exact"]) if c["seen_exact"] else "n/a",
                         pct(c["unseen_exact"]) if c["unseen_exact"] else "n/a",
                         c["mean_sim"]]); r += 1
        r += 1
        ws.cell(row=r, column=1, value="Saturation (McNemar between adjacent k, Holm-Bonferroni corrected)").font = Font(italic=True); r += 1
        _header(ws, r, ["Comparison", "Verdict (Holm-corrected)"]); r += 1
        for k, v in rq3["saturation"].items():
            _row(ws, r, [k, sig_holm(v)]); r += 1
        r += 1

    # RQ4 perception-noise simulation
    rq4 = SUMMARY.get("rq4")
    if rq4:
        ws.cell(row=r, column=1, value="RQ4 — Grasp-success vs perception noise (GEOMETRIC SIMULATION — "
                "no camera, no physics engine)").font = H2; r += 1
        ws.cell(row=r, column=1, value=rq4["caveat"]).font = Font(italic=True, color="C00000")
        ws.row_dimensions[r].height = 30; r += 1
        ws.cell(row=r, column=1,
                value=f"Tolerance basis: {rq4['tolerance_basis']} (margin shown = "
                      f"{rq4['default_compliance_margin_mm']}mm unless noted)").font = Font(italic=True)
        r += 1
        _header(ws, r, ["Pose-noise sigma (mm)", "Grasp-success rate (95% CI)"]); r += 1
        for sigma, ci in rq4["per_sigma_pooled"].items():
            _row(ws, r, [sigma, pct(ci)]); r += 1
        r += 1
        ws.cell(row=r, column=1,
                value=f"Crossover (success-rate CI upper bound drops below 50%): sigma = "
                      f"{rq4['crossover_sigma_mm']}mm — i.e. once simulated pose error exceeds this, "
                      f"the system can no longer be assumed to grasp reliably without real perception "
                      f"feedback or a wider gripper-compliance margin.").font = Font(italic=True)
        r += 2
        ws.cell(row=r, column=1, value="Sensitivity to the compliance-margin assumption (at 15mm pose noise)").font = Font(italic=True); r += 1
        _header(ws, r, ["Compliance margin (mm)", "Grasp-success rate (95% CI)"]); r += 1
        for m, ci in rq4["margin_sensitivity_at_15mm_noise"].items():
            _row(ws, r, [m, pct(ci)]); r += 1


# ---------------------------------------------------------------- Tab 4
def tab_labels(wb):
    ws = wb.create_sheet("Label Validation")
    _widths(ws, {"A": 30, "B": 34, "C": 34, "D": 8, "E": 16, "F": 14, "G": 12, "H": 10})
    ws["A1"] = "Label Validation — sample for human review"; ws["A1"].font = H1
    ws["A2"] = ("Fill 'Human correct?' (Y/N) for these rows, then compute Cohen's kappa "
                "vs the auto 'Exact' column. References flagged needs_review must be checked first.")
    ws["A2"].alignment = WRAP
    refs, vocab = load_refs(), load_vocab()
    # Was `load_json(_latest("rq1_results_*.json"))` — the LEGACY results/ path,
    # which no longer exists since the move to run_fast.py + results_fast/*.jsonl.
    # _latest() returned None, load_json(None) returned [], and this tab silently
    # rendered zero sample rows. That is why no human kappa was ever computed:
    # not because column H went unfilled, but because there was nothing to fill.
    # load_rq(1) prefers results_fast/rq1.jsonl and still falls back to legacy.
    rows = drop_fallbacks(enrich(load_rq(1), refs, vocab))
    # take a stratified-ish sample: first occurrence of each command, prefer LLM configs
    seen = {}
    for r in rows:
        if r.get("configuration") in ("LO", "LR", "FS") and r["command"] not in seen and r.get("planned_skills"):
            seen[r["command"]] = r
    sample = list(seen.values())[:30]
    if not sample:
        raise SystemExit(
            "Label Validation tab would be empty — no usable RQ1 rows found. "
            "Run `python run_fast.py --rq all` first. (Refusing to write a "
            "silently empty human-review tab; that is the failure mode this "
            "check exists to prevent.)"
        )
    # Human ratings are the one input here that cannot be regenerated. This
    # workbook is gitignored and rebuilt from scratch on every run, so ratings
    # that live only in column H are destroyed by a routine `build_workbook`.
    # They are therefore kept in eval/human_ratings.json (committed) and
    # restored into column H on every rebuild.
    human = {}
    hr_path = HERE / "human_ratings.json"
    if hr_path.exists():
        human = json.loads(hr_path.read_text(encoding="utf-8")).get("ratings", {})

    _header(ws, 4, ["Command", "Reference plan", "Predicted plan", "Exact",
                    "Failure mode", "Safety label", "Needs review", "Human correct?"])
    rr = 5
    for r in sample:
        ref = refs.get(r["command"], {})
        acc = ref.get("acceptable_reference_plans", [])
        ref_str = " ; ".join("→".join(f"{s['name']}({s.get('params',{}).get('target','')})"
                                      for s in p) for p in acc[:1]) or "(refuse)"
        pred_str = "→".join(f"{s['name']}({s.get('params',{}).get('target','')})"
                            for s in r.get("planned_skills", []))
        _row(ws, rr, [r["command"], ref_str, pred_str, "Y" if r["exact"] else "N",
                      r["failure_mode"] or "-", r["safety_label"],
                      "yes" if ref.get("needs_human_review") else "no",
                      human.get(r["command"], "?")])
        ws.row_dimensions[rr].height = 26
        rr += 1
    n_restored = sum(1 for r in sample if r["command"] in human)
    note = (f"Cohen's kappa (Human vs Auto Exact): run `python -m eval.compute_kappa`. "
            f"{n_restored}/{len(sample)} human ratings restored from human_ratings.json — "
            f"edit column H here, then mirror the change back into that file so it survives "
            f"the next rebuild."
            if n_restored else
            "Cohen's kappa (Human vs Auto Exact): __ (fill column H, then save the ratings "
            "into eval/human_ratings.json — column H alone is wiped on the next rebuild)")
    ws.cell(row=rr + 1, column=1, value=note).font = Font(italic=True, bold=True)


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    tab_exec(wb)
    tab_reco(wb)
    tab_analysis(wb)
    tab_labels(wb)
    out = HERE / "Result_robot.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
